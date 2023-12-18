import struct
import sys
import os
from dataclasses import dataclass, field
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Side, Border, Alignment


@dataclass
class PHDRRecord:
    format: str = field(init=False, default="<20sHHHIII")
    preset_name: str
    preset: int
    bank: int
    bag_ndx: int
    # the following fields are not used as per the spec 2.4
    library: int = 0
    genre: int = 0
    morphology: int = 0

    def __init__(self, preset_name: bytearray, preset, bank, bag_ndx, library, genre, morpholofy):
        self.preset_name = preset_name.decode("utf-8").strip(chr(0))
        self.preset = preset
        self.bank = bank
        self.bag_ndx = bag_ndx


@dataclass
class XBAGRecord:
    format: str = field(init=False, default="<HH")
    gen_ndx: int
    mod_ndx: int


@dataclass
class XMODRecord:
    format: str = field(init=False, default="<HHHHH")
    mod_src_oper: int
    mod_dest_oper: int
    mod_amount: int
    mod_arm_src_oper: int
    mod_trans_oper: int


@dataclass
class XGENRecord:
    format: str = field(init=False, default="<HH")
    gen_oper: int
    gen_amount: int


@dataclass
class INSTRecord:
    format: str = field(init=False, default="<20sH")
    inst_name: str
    inst_bag_ndx: int

    def __init__(self, inst_name: bytearray, inst_bag_ndx):
        self.inst_name = inst_name.decode("utf-8").strip(chr(0))
        self.inst_bag_ndx = inst_bag_ndx


@dataclass
class SHDRRecord:
    format: str = field(init=False, default="<20sIIIIIBbHH")
    sample_name: str
    start: int
    end: int
    start_loop: int
    end_loop: int
    sample_rate: int
    original_pitch: int
    pitch_correction: int
    sample_link: int
    sample_type: int

    def __init__(self, sample_name: bytearray, start, end, start_loop, end_loop, sample_rate, original_pitch,
                 pitch_correction, sample_link, sample_type):
        self.sample_name = sample_name.decode("utf-8").strip(chr(0))
        self.start = start
        self.end = end
        self.start_loop = start_loop
        self.end_loop = end_loop
        self.sample_rate = sample_rate
        self.original_pitch = original_pitch
        self.pitch_correction = pitch_correction
        self.sample_link = sample_link
        self.sample_type = sample_type


class ChunkParser:
    def __init__(self, record_class):
        self.rec_class = record_class
        self.rec_size = struct.calcsize(self.rec_class.format)
        self.records = list()

    def _parse_record(self, data, cp):
        fields = struct.unpack(self.rec_class.format, data[cp:cp + self.rec_size])
        rec = self.rec_class(*fields)
        return rec

    def parse(self, data, size):
        total_recs = int(size / self.rec_size) - 1  # ignore the terminal record
        cp = 0
        for _ in range(total_recs):
            rec = self._parse_record(data, cp)
            self.records.append(rec)
            cp += self.rec_size


@dataclass
class BankEntry:
    preset: int
    name: str
    bag_ndx: int
    zones: int
    instruments: list


GENERATORS = {0: "startAddrsOffset",
              1: "endAddrsOffset",
              2: "startloopAddrsOffset",
              3: "endloopAddrsOffset",
              4: "startAddrsCoarseOffset",
              5: "modLfoToPitch",
              6: "vibLfoToPitch",
              7: "modEnvToPitch",
              8: "initialFilterFc",
              9: "initialFilterQ",
              10: "modLfoToFilterFc",
              11: "modEnvToFilterFc",
              12: "endAddrsCoarseOffset",
              13: "modLfoToVolume",
              14: "unused1",
              15: "chorusEffectsSend",
              16: "reverbEffectsSend",
              17: "pan",
              18: "unused2",
              19: "unused3",
              20: "unused4",
              21: "delayModLFO",
              22: "freqModLFO",
              23: "delayVibLFO",
              24: "freqVibLFO",
              25: "delayModEnv",
              26: "attackModEnv",
              27: "holdModEnv",
              28: "decayModEnv",
              29: "sustainModEnv",
              30: "releaseModEnv",
              31: "keynumToModEnvHold",
              32: "keynumToModEnvDecay",
              33: "delayVolEnv",
              34: "attackVolEnv",
              35: "holdVolEnv",
              36: "decayVolEnv",
              37: "sustainVolEnv",
              38: "releaseVolEnv",
              39: "keynumToVolEnvHold",
              40: "keynumToVolEnvDecay",
              41: "instrument",
              42: "reserved1",
              43: "keyRange",
              44: "velRange",
              45: "startloopAddrsCoarseOffset",
              46: "keynum",
              47: "velocity",
              48: "initialAttenuation",
              49: "reserved2",
              50: "endloopAddrsCoarseOffset",
              51: "coarseTune",
              52: "fineTune",
              53: "sampleID",
              54: "sampleModes",
              55: "reserved3",
              56: "scaleTuning",
              57: "exclusiveClass",
              58: "overridingRootKey",
              59: "unused5",
              60: "endOper",
              }

RANGE_GENS = (43, 44)
GEN_INSTRUMENT = 41
GEN_SAMPLE_ID = 53


def chunk_id(id_str):
    id_ = 0
    for j in range(4):
        id_ |= (ord(id_str[j])) << (j * 8)
    return id_


def check_id(act_id_num, exp_id_str):
    if act_id_num != chunk_id(exp_id_str):
        print(f"File format error: expected ID '{exp_id_str}' not found")
        return False
    return True


def parse_info_list_chunk(data, cp):
    def parse_sub_chunk(data_, cp, id_):
        header_ = struct.unpack("II", data_[cp:cp+8])
        id_num = chunk_id(id_)
        if header_[0] != id_num:
            return None, None

        size_ = header_[1]
        if id_ in ["ifil", "iver"]:
            v = data[cp+8:cp+8+size_]
            maj_ver = (v[1] << 8) + v[0]
            min_ver = (v[3] << 8) + v[2]
            return f"{maj_ver}.{min_ver}", size_+8
        else:
            s = ""
            for j in range(cp+8, cp+size_+7):
                s += chr(data_[j])
            return s.strip(chr(0)), size_+8

    # ifil-ck, isng-ck, INAM-ck etc.
    chunks = {"ifil": ("Version", True),
              "isng": ("Target Sound Engine", True),
              "INAM": ("Sound Font Bank Name", True),
              "irom": ("ROM", False),
              "iver": ("ROM Revision", False),
              "ICRD": ("Date of Creation of the Bank", False),
              "IENG": ("Sound Designers and Engineers for the Bank", False),
              "IPRD": ("Product for which the Bank was intended", False),
              "ICOP": ("Copyright", False),
              "ICMT": ("Comments", False),
              "ISFT": ("SoundFont tools used to create and alter the bank", False)}

    info = list()
    for id_, name in chunks.items():
        data_str, size = parse_sub_chunk(data, cp, id_)
        if data_str is None:
            if name[1]:
                print(f"File does not have have RIFF format. {name[0]} sub-chunk is not present")
                return
            else:
                # the sub-chunk is optional
                continue
        tup = (name[0], data_str)
        info.append(tup)
        cp += size

    return info


def parse_chunk(f, rec_class, chunk_id: str):
    data = f.read(8)
    header = struct.unpack("II", data)

    if not check_id(header[0], chunk_id):
        sys.exit(1)

    size = header[1]
    data = f.read(size)

    chunk = ChunkParser(rec_class)
    chunk.parse(data, size)

    return chunk


def parse_file(fpath):
    file_size = os.path.getsize(fpath)

    f = open(fpath, "rb")
    data = f.read(1024)

    # parse first 24 bytes
    header = struct.unpack("IIIIII", data[:24])

    # SFBK-form
    if not check_id(header[0], "RIFF"):
        sys.exit(1)

    riff_size = header[1]
    if riff_size + 8 != file_size:
        print("File is incomplete")
        sys.exit(1)

    if not check_id(header[2], "sfbk"):
        sys.exit(1)

    if not check_id(header[3], "LIST"):
        sys.exit(1)

    size = header[4]     # includes 'INFO' ID

    if not check_id(header[5], "INFO"):
        sys.exit(1)

    info = parse_info_list_chunk(data, 24)
    curr_pos = size + 20    # exclude 'INFO' ID

    # SDTI chunk (samples)
    header = struct.unpack("II", data[curr_pos:curr_pos+8])
    if not check_id(header[0], "LIST"):
        sys.exit(1)

    size = header[1]
    curr_pos += 8

    # skip samples
    curr_pos += size
    f.seek(curr_pos, 0)

    # PDTA chunk
    data = f.read(12)
    header = struct.unpack("III", data)
    if not check_id(header[0], "LIST"):
        sys.exit(1)

    # size = header[1]
    if not check_id(header[2], "pdta"):
        sys.exit(1)

    # PDHR, PBAG etc. chunks
    phdr = parse_chunk(f, PHDRRecord, "phdr")
    pbag = parse_chunk(f, XBAGRecord, "pbag")
    pmod = parse_chunk(f, XMODRecord, "pmod")
    pgen = parse_chunk(f, XGENRecord, "pgen")
    inst = parse_chunk(f, INSTRecord, "inst")
    ibag = parse_chunk(f, XBAGRecord, "ibag")
    imod = parse_chunk(f, XMODRecord, "imod")
    igen = parse_chunk(f, XGENRecord, "igen")
    shdr = parse_chunk(f, SHDRRecord, "shdr")

    f.close()

    return info, phdr, pbag, pmod, pgen, inst, ibag, imod, igen, shdr


def next_bag(phdr, bag_ndx):
    res = None
    for rec in phdr.records:
        if rec.bag_ndx > bag_ndx:
            if res is None:
                res = rec.bag_ndx
            else:
                if rec.bag_ndx < res:
                    res = rec.bag_ndx
    return res


def next_gen(pbag, gen_ndx):
    res = None
    for rec in pbag.records:
        if rec.gen_ndx > gen_ndx:
            if res is None:
                res = rec.gen_ndx
            else:
                if rec.gen_ndx < res:
                    res = rec.gen_ndx
    return res


def process(phdr, pbag, pmod, pgen, inst, ibag, imod, igen, shdr):
    # sort recs
    phdr.records.sort(key=lambda r: r.bank)

    banks = {}
    for rec in phdr.records:
        zones = 0
        next = next_bag(phdr, rec.bag_ndx)
        if next is None:
            zones = len(pbag.records) - rec.bag_ndx
        else:
            zones = next - rec.bag_ndx

        b = rec.bank
        if b not in banks.keys():
            banks[b] = list()

        instruments = list()
        for bag_idx in range(rec.bag_ndx, rec.bag_ndx+zones):
            bag = pbag.records[bag_idx]

            gens = 0
            next = next_gen(pbag, bag.gen_ndx)
            if next is None:
                gens = len(pgen.records) - bag.gen_ndx
            else:
                gens = next - bag.gen_ndx

            for gen_idx in range(bag.gen_ndx, bag.gen_ndx+gens):
                gen = pgen.records[gen_idx]
                # print(bag_idx, gen_idx, gen)
                if gen.gen_oper == 41:
                    # instrument
                    ins = inst.records[gen.gen_amount]
                    if ins.inst_name not in instruments:
                        instruments.append(ins.inst_name)

        banks[b].append(BankEntry(preset=rec.preset, name=rec.preset_name, bag_ndx=rec.bag_ndx, zones=zones, instruments=instruments))

    return banks


def unpack_amount(gen, amount):
    if gen in RANGE_GENS:
        return f"{amount & 0xff} - {(amount & 0xff00) >> 8}"

    return amount


def write_xlsx(fpath, info, banks, phdr, pbag, pmod, pgen, inst, ibag, imod, igen, shdr):
    def apply_style(ws, freeze=True):
        bk_color_1 = PatternFill(start_color='2f2f2f', end_color='2f2f2f', fill_type="solid")
        bk_color_2 = PatternFill(start_color='1f1f1f', end_color='1f1f1f', fill_type="solid")
        font = Font(name="Consolas", size=9, color="c0c0c0")
        alignment = Alignment(horizontal='left', vertical='center')
        alignment_wrap = Alignment(horizontal='left', vertical='center', wrapText=True)
        thin = Side(border_style="thin", color="707070")
        bk_1 = False
        next_is_comment = False
        range = f"A1:Z{ws.max_row+100}"
        for row in ws[range]:
            bk = bk_color_1 if bk_1 else bk_color_2
            bk_1 = not bk_1
            for cell in row:
                cell.font = font

                cell.alignment = alignment
                if cell.value and cell.value == "Comments":
                    next_is_comment = True
                elif next_is_comment:
                    cell.alignment = alignment_wrap
                    next_is_comment = False

                if cell.value is not None:
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    cell.fill = bk
                else:
                    cell.fill = bk_color_2

        if freeze:
            ws.freeze_panes = ws['A2']

        # adjust the column width
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:  # avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = max_length + 1

    wb = Workbook()

    print(f"Writing file: {fpath}...")

    print("Writing INFO tab...")
    ws = wb.active
    ws.title = "INFO"
    comment_found = ""
    for i in info:
        if i[0].startswith("Comments"):
            comment_found = i[1]
            continue
        else:
            ws.append((i[0], i[1]))
    if comment_found:
        comment_found = comment_found.replace(chr(0), "").rstrip("\n")
        ws.append(("Comments", comment_found))
    apply_style(ws, freeze=False)

    print("Writing PHDR tab...")
    wb.create_sheet("PHDR")
    ws = wb["PHDR"]
    ws.append(("achPresetName", "wPreset",  "wBank", "wPresetBagNdx", "Zones"))
    for bank, entries in banks.items():
        bank_no = bank
        entries.sort(key=lambda e: e.preset)
        for ent in entries:
            ws.append((ent.name, ent.preset, bank_no, ent.bag_ndx, ent.zones))
    apply_style(ws)

    print("Writing PBAG tab...")
    wb.create_sheet("PBAG")
    ws = wb["PBAG"]
    ws.append(("Index", "wGenNdx", "wModNdx"))
    for idx, b in enumerate(pbag.records):
        ws.append((idx, b.gen_ndx, b.mod_ndx))
    apply_style(ws)

    # TODO debug PMOD
    print("Writing PMOD tab...")
    wb.create_sheet("PMOD")
    ws = wb["PMOD"]
    ws.append(("Index", "sfModSrcOper", "sfModDestOper", "modAmount", "sfModArmSrcOper", "sfModTransOper"))
    for idx, m in enumerate(pmod.records):
        ws.append((idx, m.mod_src_oper, m.mod_dest_oper, unpack_amount(m.mod_src_oper, m.mod_amount), m.mod_arm_src_oper, m.mod_trans_oper))
    apply_style(ws)

    print("Writing PGEN tab...")
    wb.create_sheet("PGEN")
    ws = wb["PGEN"]
    ws.append(("Index", "sfGenOper", "Oper Name", "genAmount", "Inst Name"))
    for idx, g in enumerate(pgen.records):
        inst_name = inst.records[g.gen_amount].inst_name if g.gen_oper == GEN_INSTRUMENT else ""
        ws.append((idx, g.gen_oper, GENERATORS[g.gen_oper], unpack_amount(g.gen_oper, g.gen_amount), inst_name))
    apply_style(ws)

    print("Writing INST tab...")
    wb.create_sheet("INST")
    ws = wb["INST"]
    ws.append(("Index", "achInstName", "wInstBagNdx"))
    for idx, ins in enumerate(inst.records):
        ws.append((idx, ins.inst_name, ins.inst_bag_ndx))
    apply_style(ws)

    print("Writing IBAG tab...")
    wb.create_sheet("IBAG")
    ws = wb["IBAG"]
    ws.append(("Index", "wInstGenNdx", "wInstModNdx"))
    for idx, b in enumerate(ibag.records):
        ws.append((idx, b.gen_ndx, b.mod_ndx))
    apply_style(ws)

    # TODO debug IMOD
    print("Writing IMOD tab...")
    wb.create_sheet("IMOD")
    ws = wb["IMOD"]
    ws.append(("Index", "sfModSrcOper", "sfModDestOper", "modAmount", "sfModArmSrcOper", "sfModTransOper"))
    for idx, m in enumerate(imod.records):
        ws.append((idx, m.mod_src_oper, m.mod_dest_oper, unpack_amount(m.mod_src_oper, m.mod_amount), m.mod_arm_src_oper, m.mod_trans_oper))
    apply_style(ws)

    print("Writing IGEN tab...")
    wb.create_sheet("IGEN")
    ws = wb["IGEN"]
    ws.append(("Index", "sfGenOper", "Oper Name", "genAmount", "Sample Name"))
    for idx, g in enumerate(igen.records):
        sample_name = shdr.records[g.gen_amount].sample_name if g.gen_oper == GEN_SAMPLE_ID else ""
        ws.append((idx, g.gen_oper, GENERATORS[g.gen_oper], unpack_amount(g.gen_oper, g.gen_amount), sample_name))
    apply_style(ws)

    print("Writing SHDR tab...")
    wb.create_sheet("SHDR")
    ws = wb["SHDR"]
    ws.append(("Index", "achSampleName", "dwStart", "dwEnd", "dwStartloop", "dwEndloop", "dwSampleRate", "byOriginalPitch",
               "chPitchCorrection", "wSampleLink", "sfSampleType"))
    for idx, s in enumerate(shdr.records):
        try:
            ws.append((idx, s.sample_name, s.start, s.end, s.start_loop, s.end_loop, s.sample_rate, s.original_pitch, s.pitch_correction, s.sample_link, s.sample_type))
        except:
            print(s.sample_name, s.pitch_correction)
    apply_style(ws)

    wb.save(fpath)


def write_md(fpath, info, banks, phdr, pbag, pmod, pgen, inst, ibag, imod, igen, shdr):
    print(f"\nWriting file: {fpath}...\n")

    with open(fpath, "wt") as f:
        f.write(f"# File {fpath}\n\n")

        comment_found = ""
        for i in info:
            if i[0].startswith("Comments"):
                comment_found = i[1]
                continue
            else:
                f.write(f"- {i[0]}: `{i[1]}`\n")
        if comment_found:
            f.write("### Comments\n\n")
            comment_found = comment_found.replace(chr(0), "").rstrip("\n")
            f.write(f"{comment_found}\n")

        f.write(f"\n## Presets\n")

        f.write(f"- Banks: `{len(banks)}`\n")
        f.write(f"- Presets: `{len(phdr.records)}`\n")
        f.write(f"- Instruments: `{len(inst.records)}`\n")
        f.write(f"- Samples: `{len(shdr.records)}`\n\n")

        f.write("| Bank | Preset | Preset Name        | Instruments |\n")
        f.write("|:----:|:------:|--------------------|-------------|\n")
        for bank, entries in banks.items():
            bank_no = bank
            entries.sort(key=lambda e: e.preset)
            for ent in entries:
                f.write(f"|{bank_no}|{ent.preset}|{ent.name}|{", ".join(ent.instruments)}\n")
                bank_no = " "

        f.write(f"\n## Instruments\n")

        f.write("| Instrument | Instrument Name    |\n")
        f.write("|:----------:|--------------------|\n")
        for idx, ins in enumerate(inst.records):
            f.write(f"|{idx}|{ins.inst_name}|\n")


def main(fpath):
    print(f"Reading file: {fpath}...\n")
    if not os.path.exists(fpath) or not os.path.isfile(fpath):
        print("File not found.")
        return

    info, phdr, pbag, pmod, pgen, inst, ibag, imod, igen, shdr = parse_file(fpath)

    banks = process(phdr, pbag, pmod, pgen, inst, ibag, imod, igen, shdr)

    # save the contents
    out_fpath = fpath.replace(".sf2", ".xlsx")
    write_xlsx(out_fpath, info, banks, phdr, pbag, pmod, pgen, inst, ibag, imod, igen, shdr)
    out_fpath = fpath.replace(".sf2", ".md")
    write_md(out_fpath, info, banks, phdr, pbag, pmod, pgen, inst, ibag, imod, igen, shdr)

    print("Done.")


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: python sf2-contents.py <file.sf2>")

    main(sys.argv[1])
